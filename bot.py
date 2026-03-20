import os, re, random, logging
from datetime import datetime, timezone, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (ApplicationBuilder, MessageHandler, CommandHandler,
                          ConversationHandler, CallbackQueryHandler,
                          filters, ContextTypes)
from supabase import create_client

TOKEN          = os.environ["TELEGRAM_TOKEN"]
SUPABASE_URL   = os.environ["SUPABASE_URL"]
SUPABASE_KEY   = os.environ["SUPABASE_KEY"]
DASHBOARD_URL  = os.environ.get("DASHBOARD_URL", "https://tu-panel.vercel.app")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
logging.basicConfig(level=logging.INFO)
ARG = timezone(timedelta(hours=-3))

# ── Estados ──────────────────────────────────────────────────
(
    REG_ROL, REG_NOMBRE,
    ADD_OP_NOMBRE, ADD_OP_SOY_YO,
    ADD_CLI_NOMBRE, ADD_CLI_SOY_YO,
    INGRESAR_CODIGO, CONFIRMAR_NOMBRE, CORREGIR_NOMBRE,
    DESC_CLIENTE, DESC_CAMPO, DESC_LOTE, DESC_GRANO_OTRO,
    DESC_TIPO_DESTINO,
    DESC_CAMION_CHASIS, DESC_CAMION_ACOPLADO, DESC_CAMION_CAPACIDAD,
    DESC_KG,
    NUEVO_LOTE_NOMBRE, NUEVO_CAMPO_NOMBRE, NUEVO_CLIENTE_NOMBRE,
) = range(21)

GRANOS = ["Trigo", "Soja", "Maíz", "Girasol", "Sorgo"]

def generar_codigo():
    return str(random.randint(1000, 9999))


def ahora():
    return datetime.now(ARG)

# ── BD helpers ───────────────────────────────────────────────
def get_contratista(telegram_id: str):
    r = supabase.table("contratistas").select("*").eq("telegram_id", telegram_id).execute()
    return r.data[0] if r.data else None

def get_usuario(telegram_id: str):
    r = supabase.table("usuarios").select("*, contratistas(nombre,apellido)").eq("telegram_id", telegram_id).execute()
    return r.data[0] if r.data else None

def get_cliente_by_telegram(telegram_id: str):
    r = supabase.table("clientes").select("*, contratistas(nombre,apellido)").eq("telegram_id", telegram_id).execute()
    return r.data[0] if r.data else None

def get_operarios(contratista_id: int):
    r = supabase.table("usuarios").select("*").eq("contratista_id", contratista_id).eq("rol", "operario").execute()
    return r.data or []

def get_clientes(contratista_id: int):
    r = supabase.table("clientes").select("*").eq("contratista_id", contratista_id).execute()
    return r.data or []

def es_operario(telegram_id: str):
    r = supabase.table("usuarios").select("id").eq("telegram_id", telegram_id).eq("rol", "operario").execute()
    return bool(r.data)
  
def get_contratista_id_de_usuario(telegram_id: str):
    cont = get_contratista(telegram_id)
    if cont: return cont["id"]
    usr = get_usuario(telegram_id)
    if usr: return usr["contratista_id"]
    return None

def get_sesion(chat_id: str):
    r = supabase.table("sesion_activa").select(
        "*, clientes(nombre,apellido), campos(nombre), lotes(nombre,grano)"
    ).eq("chat_id", chat_id).execute()
    return r.data[0] if r.data else None

def get_sesion_por_contratista(contratista_id: int):
    r = supabase.table("sesion_activa").select(
        "*, clientes(nombre,apellido), campos(nombre), lotes(nombre,grano)"
    ).eq("contratista_id", contratista_id).order("iniciada_at", desc=True).limit(1).execute()
    return r.data[0] if r.data else None

def get_campos(cliente_id: int):
    r = supabase.table("campos").select("*").eq("cliente_id", cliente_id).order("nombre").execute()
    return r.data or []

def get_lotes(campo_id: int):
    r = supabase.table("lotes").select("*").eq("campo_id", campo_id).order("nombre").execute()
    return r.data or []

def get_camiones_abiertos(contratista_id: int):
    r = supabase.table("camiones").select("*").eq("contratista_id", contratista_id).eq("cerrado", False).execute()
    camiones = []
    for c in r.data or []:
        acum = kg_acumulado_camion(c["id"])
        camiones.append({**c, "acumulado": acum})
    return camiones

def get_silobolsas_abiertos(contratista_id: int):
    r = (supabase.table("silobolsas")
         .select("*, lotes(nombre, grano, campos(nombre))")
         .eq("cerrado", False)
         .execute())
    silos = []
    for s in r.data or []:
        lote = s.get("lotes") or {}
        campo = lote.get("campos") or {}
        if not lote: continue
        acum = kg_acumulado_silo(s["id"])
        silos.append({**s, "acumulado": acum, "lote_nombre": lote.get("nombre",""), "campo_nombre": campo.get("nombre",""), "grano": lote.get("grano","")})
    return silos

def kg_acumulado_camion(camion_id: int):
    r = supabase.table("descargas").select("kg").eq("camion_id", camion_id).execute()
    return sum(float(x["kg"]) for x in r.data) if r.data else 0

def kg_acumulado_silo(silo_id: int):
    r = supabase.table("descargas").select("kg").eq("silobolsa_id", silo_id).execute()
    return sum(float(x["kg"]) for x in r.data) if r.data else 0

def barra(actual, total):
    if not total: return ""
    pct    = min(actual / total, 1.0)
    llenos = int(pct * 20)
    return "█" * llenos + "░" * (20 - llenos) + f" {pct*100:.0f}%"

def tiene_descargas_operario(operario_id: int) -> bool:
    r = supabase.table("descargas").select("id").eq("operario_id", operario_id).limit(1).execute()
    return bool(r.data)

def tiene_descargas_cliente(cliente_id: int) -> bool:
    r = supabase.table("descargas").select("id").eq("cliente_id", cliente_id).limit(1).execute()
    return bool(r.data)

def parsear_kg(texto: str):
    t = texto.upper()
    m = re.search(r'(\d[\d\.,]*)\s*(?:KGS?|KL|KILO|TON(?:ES|S)?)?', t)
    if not m: return None
    val = m.group(1).replace('.','').replace(',','.')
    kg  = float(val)
    if re.search(r'TON', t): kg *= 1000
    return kg if kg > 0 else None

def parsear_patente(texto: str):
    t = texto.upper().strip()
    m = re.search(r'\b([A-Z]{2,3}\s?\d{3}\s?[A-Z]{0,2})\b', t)
    return re.sub(r'\s+', '', m.group(1)) if m else re.sub(r'\s+', '', t)

def get_descargas_lote(lote_id: int):
    r = supabase.table("descargas").select(
        "*, clientes(nombre,apellido), campos(nombre), lotes(nombre,grano), "
        "camiones(patente_chasis,patente_acoplado,cerrado), silobolsas(numero,cerrado), usuarios(nombre)"
    ).eq("lote_id", lote_id).order("created_at").execute()
    return r.data or []

def get_descargas_contratista(contratista_id: int):
    r = supabase.table("descargas").select(
        "*, clientes(nombre,apellido), campos(nombre), lotes(nombre,grano), "
        "camiones(patente_chasis,patente_acoplado,cerrado), silobolsas(numero,cerrado), usuarios(nombre)"
    ).eq("contratista_id", contratista_id).order("created_at").execute()
    return r.data or []

def generar_excel(descargas: list) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Descargas"

    headers = ["Fecha", "Cliente", "Campo", "Lote", "Grano", "Tipo", "Destino", "Kg", "Operario"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    for d in descargas:
        cliente = d.get("clientes")   or {}
        campo   = d.get("campos")     or {}
        lote    = d.get("lotes")      or {}
        camion  = d.get("camiones")   or {}
        silo    = d.get("silobolsas") or {}
        op      = d.get("usuarios")   or {}

        fecha       = (d.get("created_at") or "")[:10]
        cliente_str = f"{cliente.get('nombre','')} {cliente.get('apellido','')}".strip()
        tipo        = "Camión" if d.get("destino") == "camion" else "Silobolsa"
        destino     = (
            f"{camion.get('patente_chasis','')} / {camion.get('patente_acoplado','')}"
            if d.get("destino") == "camion"
            else f"Silo #{silo.get('numero','')}"
        )
        ws.append([
            fecha, cliente_str, campo.get("nombre", ""), lote.get("nombre", ""),
            lote.get("grano", ""), tipo, destino, float(d.get("kg", 0)), op.get("nombre", "")
        ])

    # Ajustar anchos de columna
    col_widths = [12, 20, 18, 18, 10, 10, 24, 10, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Teclados ─────────────────────────────────────────────────
def teclado_menu_contratista(contratista_id: int, telegram_id: str):
    operarios = get_operarios(contratista_id)
    clientes  = get_clientes(contratista_id)
    botones   = [
        [InlineKeyboardButton(f"👷 Mis operarios ({len(operarios)})", callback_data="cont_ver_op")],
        [InlineKeyboardButton(f"👤 Mis clientes ({len(clientes)})",   callback_data="cont_ver_cli")],
        [InlineKeyboardButton("🚛 Mis camiones",  callback_data="op_camiones"),
         InlineKeyboardButton("🌾 Mis silobolsas", callback_data="op_silos")],
        [InlineKeyboardButton("📊 Mis descargas",                     callback_data="cont_ver_desc")],
    ]
    if es_operario(telegram_id):
        botones.append([InlineKeyboardButton("📦 Agregar descarga", callback_data="op_descarga")])
    return InlineKeyboardMarkup(botones)

def teclado_menu_operario():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📦 Agregar descarga", callback_data="op_descarga")],
        [InlineKeyboardButton("🚛 Mis camiones",     callback_data="op_camiones")],
        [InlineKeyboardButton("🌾 Mis silobolsas",   callback_data="op_silos")],
        [InlineKeyboardButton("📊 Mis descargas",    callback_data="cont_ver_desc")],
    ])

def teclado_menu_cliente():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 Mis descargas", callback_data="cli_ver_desc")],
    ])

def teclado_roles():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🏭 Contratista / Dueño del equipo", callback_data="rol_contratista")],
        [InlineKeyboardButton("🌾 Operario de tolva",              callback_data="rol_operario")],
        [InlineKeyboardButton("👤 Cliente / Dueño de granos",      callback_data="rol_cliente")],
    ])

def btn_cancelar(cb="op_cancelar"):
    return InlineKeyboardButton("❌ Cancelar", callback_data=cb)

# ── /start ───────────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid  = str(update.effective_user.id)
    cont = get_contratista(uid)
    if cont:
        await update.message.reply_text(
            f"Hola {cont['nombre']}! ¿Qué querés hacer?",
            reply_markup=teclado_menu_contratista(cont["id"], uid)
        )
        return ConversationHandler.END
    usr = get_usuario(uid)
    if usr:
        await update.message.reply_text(
            f"Hola {usr['nombre']}! ¿Qué querés hacer?",
            reply_markup=teclado_menu_operario()
        )
        return ConversationHandler.END
    cli = get_cliente_by_telegram(uid)
    if cli:
        cont_nombre = (cli.get("contratistas") or {}).get("nombre", "")
        await update.message.reply_text(
            f"Hola {cli['nombre']} {cli['apellido']}! Sos cliente de *{cont_nombre}*.\n¿Qué querés hacer?",
            parse_mode="Markdown",
            reply_markup=teclado_menu_cliente()
        )
        return ConversationHandler.END
    await update.message.reply_text(
        "Hola! Bienvenido al sistema de tolvas.\n\n¿Quién sos?",
        reply_markup=teclado_roles()
    )
    return REG_ROL

# ── Registro ─────────────────────────────────────────────────
async def elegir_rol(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    rol   = query.data.replace("rol_", "")
    if rol == "contratista":
        await query.edit_message_text("¿Cuál es tu nombre y apellido?")
        return REG_NOMBRE
    await query.edit_message_text("Ingresá el código de acceso que te dio tu contratista:")
    return INGRESAR_CODIGO

async def recibir_nombre_contratista(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto    = update.message.text.strip()
    partes   = texto.split()
    nombre   = partes[0]
    apellido = " ".join(partes[1:]) if len(partes) > 1 else ""
    uid      = str(update.effective_user.id)
    supabase.table("contratistas").insert({
        "nombre": nombre, "apellido": apellido, "telegram_id": uid,
    }).execute()
    cont = get_contratista(uid)

    await update.message.reply_text(
        f"✅ Bienvenido *{nombre} {apellido}*! Quedaste registrado como contratista.\n\n"
        f"📱 *Tu ID de Telegram:* `{uid}`\n\n"
        f"Para acceder al panel web entrá a:\n{DASHBOARD_URL}/register\n"
        f"y usá tu ID de Telegram para crear tu cuenta.",
        parse_mode="Markdown",
        reply_markup=teclado_menu_contratista(cont["id"], uid)
    )
    return ConversationHandler.END

async def ingresar_codigo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    codigo = update.message.text.strip()
    uid    = str(update.effective_user.id)
    r = supabase.table("usuarios").select("*").eq("codigo_acceso", codigo).is_("telegram_id", "null").execute()
    if r.data:
        usuario = r.data[0]
        context.user_data.update({
            "codigo_encontrado_id": usuario["id"],
            "codigo_encontrado_tipo": "operario",
            "codigo_encontrado_nombre": usuario["nombre"],
        })
        teclado = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Sí, está bien",      callback_data="nombre_ok")],
            [InlineKeyboardButton("✏️ Corregir mi nombre", callback_data="nombre_corregir")],
        ])
        await update.message.reply_text(
            f"✅ Código correcto. Tu contratista te registró como:\n\n*{usuario['nombre']}*\n\n¿Es correcto?",
            parse_mode="Markdown", reply_markup=teclado
        )
        return CONFIRMAR_NOMBRE
    r = supabase.table("clientes").select("*").eq("codigo_acceso", codigo).is_("telegram_id", "null").execute()
    if r.data:
        cliente = r.data[0]
        context.user_data.update({
            "codigo_encontrado_id": cliente["id"],
            "codigo_encontrado_tipo": "cliente",
            "codigo_encontrado_nombre": f"{cliente['nombre']} {cliente['apellido']}",
        })
        teclado = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Sí, está bien",      callback_data="nombre_ok")],
            [InlineKeyboardButton("✏️ Corregir mi nombre", callback_data="nombre_corregir")],
        ])
        await update.message.reply_text(
            f"✅ Código correcto. Tu contratista te registró como:\n\n*{cliente['nombre']} {cliente['apellido']}*\n\n¿Es correcto?",
            parse_mode="Markdown", reply_markup=teclado
        )
        return CONFIRMAR_NOMBRE
    await update.message.reply_text(
        "Código incorrecto o ya fue usado.\nPedile a tu contratista el código correcto e intentá de nuevo:"
    )
    return INGRESAR_CODIGO

async def confirmar_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "nombre_corregir":
        await query.edit_message_text("¿Cuál es tu nombre y apellido correcto?")
        return CORREGIR_NOMBRE
    await _vincular_usuario(str(update.effective_user.id), context, None, query)
    return ConversationHandler.END

async def corregir_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nombre_nuevo = update.message.text.strip()
    await _vincular_usuario(str(update.effective_user.id), context, nombre_nuevo, None, update.message)
    return ConversationHandler.END

async def _vincular_usuario(uid, context, nombre_nuevo, query=None, msg=None):
    tipo   = context.user_data["codigo_encontrado_tipo"]
    rec_id = context.user_data["codigo_encontrado_id"]
    if tipo == "operario":
        update_data = {"telegram_id": uid, "codigo_acceso": None}
        if nombre_nuevo: update_data["nombre"] = nombre_nuevo.split()[0]
        supabase.table("usuarios").update(update_data).eq("id", rec_id).execute()
        r    = supabase.table("usuarios").select("nombre").eq("id", rec_id).execute()
        nombre_final = r.data[0]["nombre"] if r.data else ""
        texto   = f"✅ Bienvenido *{nombre_final}*! Ya tenés acceso como operario."
        teclado = teclado_menu_operario()
    else:
        update_data = {"telegram_id": uid, "codigo_acceso": None}
        if nombre_nuevo:
            partes = nombre_nuevo.split()
            update_data["nombre"]   = partes[0]
            update_data["apellido"] = " ".join(partes[1:]) if len(partes) > 1 else ""
        supabase.table("clientes").update(update_data).eq("id", rec_id).execute()
        r = supabase.table("clientes").select("nombre,apellido").eq("id", rec_id).execute()
        c = r.data[0] if r.data else {}
        nombre_final = f"{c.get('nombre','')} {c.get('apellido','')}".strip()
        texto = (
            f"✅ Bienvenido *{nombre_final}*! Ya tenés acceso como cliente.\n\n"
            f"📱 *Tu ID de Telegram:* `{uid}`\n\n"
            f"Para acceder al panel web entrá a:\n{DASHBOARD_URL}/register\n"
            f"y usá tu ID de Telegram para crear tu cuenta.\n"
            f"_Podés crear múltiples cuentas con el mismo ID si tenés empleados que también necesitan acceso._"
        )
        teclado = None
    if query:
        await query.edit_message_text(texto, parse_mode="Markdown", reply_markup=teclado)
    elif msg:
        await msg.reply_text(texto, parse_mode="Markdown", reply_markup=teclado)

# ── Menú contratista ─────────────────────────────────────────
async def menu_contratista_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query  = update.callback_query
    await query.answer()
    uid    = str(update.effective_user.id)
    cont   = get_contratista(uid)
    accion = query.data

    # Acciones exclusivas para clientes (no requieren cont)
    es_accion_cli = accion in ("cli_ver_desc", "cli_menu")
    # Acciones de "Mis descargas" para operarios/contratistas
    es_accion_desc = (accion == "cont_ver_desc" or accion.startswith("cont_desc_"))

    contratista_id = cont["id"] if cont else (get_contratista_id_de_usuario(uid) if es_accion_desc else None)

    if not cont and not es_accion_desc and not es_accion_cli:
        await query.edit_message_text("No se encontró tu cuenta. Escribí /start.")
        return ConversationHandler.END

    if es_accion_desc and not cont and not contratista_id:
        await query.edit_message_text("No se encontró tu contratista. Escribí /start.")
        return ConversationHandler.END

    # ── Menú cliente ──────────────────────────────────────────
    if accion == "cli_menu":
        cli = get_cliente_by_telegram(uid)
        if not cli:
            await query.edit_message_text("No se encontró tu cuenta. Escribí /start.")
            return ConversationHandler.END
        cont_nombre = (cli.get("contratistas") or {}).get("nombre", "")
        await query.edit_message_text(
            f"Hola {cli['nombre']} {cli['apellido']}! Sos cliente de *{cont_nombre}*.\n¿Qué querés hacer?",
            parse_mode="Markdown", reply_markup=teclado_menu_cliente()
        )
        return ConversationHandler.END

    if accion == "cli_ver_desc":
        cli = get_cliente_by_telegram(uid)
        if not cli:
            await query.edit_message_text("No se encontró tu cuenta. Escribí /start.")
            return ConversationHandler.END
        campos  = get_campos(cli["id"])
        botones = []
        for c in campos:
            botones.append([InlineKeyboardButton(f"🌿 {c['nombre']}", callback_data=f"cont_desc_campo_{c['id']}")])
        if not campos:
            botones.append([InlineKeyboardButton("(sin campos)", callback_data="cli_menu")])
        botones.append([InlineKeyboardButton("⬅️ Volver", callback_data="cli_menu")])
        await query.edit_message_text(
            "📊 *Mis descargas*\n¿En qué campo?",
            parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones)
        )
        return ConversationHandler.END

    if accion == "cont_ver_op":
        operarios = get_operarios(cont["id"])
        botones   = []
        lineas    = ["👷 *Mis operarios*\n"] if operarios else ["No tenés operarios registrados todavía."]
        for op in operarios:
            icono = "✅" if op.get("telegram_id") else "⏳"
            lineas.append(f"{icono} {op['nombre']}")
            botones.append([InlineKeyboardButton(f"{icono} {op['nombre']}", callback_data=f"op_detalle_{op['id']}")])
        botones.append([InlineKeyboardButton("➕ Agregar operario", callback_data="cont_add_op")])
        botones.append([InlineKeyboardButton("⬅️ Volver",           callback_data="cont_volver")])
        await query.edit_message_text("\n".join(lineas), parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))
        return ConversationHandler.END

    elif accion == "cont_ver_cli":
        clientes = get_clientes(cont["id"])
        botones  = []
        lineas   = ["👤 *Mis clientes*\n"] if clientes else ["No tenés clientes registrados todavía."]
        for cli in clientes:
            icono = "✅" if cli.get("telegram_id") else "⏳"
            lineas.append(f"{icono} {cli['nombre']} {cli['apellido']}")
            botones.append([InlineKeyboardButton(f"{icono} {cli['nombre']} {cli['apellido']}", callback_data=f"cli_detalle_{cli['id']}")])
        botones.append([InlineKeyboardButton("➕ Agregar cliente", callback_data="cont_add_cli")])
        botones.append([InlineKeyboardButton("⬅️ Volver",          callback_data="cont_volver")])
        await query.edit_message_text("\n".join(lineas), parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))
        return ConversationHandler.END

    elif accion.startswith("op_detalle_"):
        op_id = int(accion.replace("op_detalle_", ""))
        r     = supabase.table("usuarios").select("*").eq("id", op_id).execute()
        if not r.data:
            await query.edit_message_text("No se encontró el operario.")
            return ConversationHandler.END
        op    = r.data[0]
        icono = "✅ Activo" if op.get("telegram_id") else "⏳ Pendiente de alta"
        botones = []
        if not op.get("telegram_id") and op.get("codigo_acceso"):
            botones.append([InlineKeyboardButton(f"🔑 Ver código: {op['codigo_acceso']}", callback_data=f"op_vercodigo_{op_id}")])
        botones.append([InlineKeyboardButton("✏️ Modificar nombre",  callback_data=f"op_editar_{op_id}")])
        botones.append([InlineKeyboardButton("🗑️ Eliminar operario", callback_data=f"op_eliminar_{op_id}")])
        botones.append([InlineKeyboardButton("⬅️ Volver",            callback_data="cont_ver_op")])
        texto_op = f"👷 *{op['nombre']}*\nEstado: {icono}"
        if op.get("telegram_id"):
            texto_op += f"\nTelegram ID: `{op['telegram_id']}`"
        elif op.get("codigo_acceso"):
            texto_op += f"\nCódigo de acceso: `{op['codigo_acceso']}`"
        await query.edit_message_text(texto_op, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))
        return ConversationHandler.END

    elif accion.startswith("cli_detalle_"):
        cli_id = int(accion.replace("cli_detalle_", ""))
        r      = supabase.table("clientes").select("*").eq("id", cli_id).execute()
        if not r.data:
            await query.edit_message_text("No se encontró el cliente.")
            return ConversationHandler.END
        cli   = r.data[0]
        icono = "✅ Activo" if cli.get("telegram_id") else "⏳ Pendiente de alta"
        botones = []
        if not cli.get("telegram_id") and cli.get("codigo_acceso"):
            botones.append([InlineKeyboardButton(f"🔑 Ver código: {cli['codigo_acceso']}", callback_data=f"cli_vercodigo_{cli_id}")])
        botones.append([InlineKeyboardButton("✏️ Modificar nombre", callback_data=f"cli_editar_{cli_id}")])
        botones.append([InlineKeyboardButton("🗑️ Eliminar cliente", callback_data=f"cli_eliminar_{cli_id}")])
        botones.append([InlineKeyboardButton("⬅️ Volver",           callback_data="cont_ver_cli")])
        texto_cli = f"👤 *{cli['nombre']} {cli['apellido']}*\nEstado: {icono}"
        if cli.get("telegram_id"):
            texto_cli += f"\nTelegram ID: `{cli['telegram_id']}`"
        elif cli.get("codigo_acceso"):
            texto_cli += f"\nCódigo de acceso: `{cli['codigo_acceso']}`"
        await query.edit_message_text(texto_cli, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))
        return ConversationHandler.END

    elif accion.startswith("op_vercodigo_"):
        op_id = int(accion.replace("op_vercodigo_", ""))
        r     = supabase.table("usuarios").select("nombre,codigo_acceso").eq("id", op_id).execute()
        if r.data: await query.answer(f"Código de {r.data[0]['nombre']}: {r.data[0]['codigo_acceso']}", show_alert=True)
        return ConversationHandler.END

    elif accion.startswith("cli_vercodigo_"):
        cli_id = int(accion.replace("cli_vercodigo_", ""))
        r      = supabase.table("clientes").select("nombre,apellido,codigo_acceso").eq("id", cli_id).execute()
        if r.data:
            c = r.data[0]
            await query.answer(f"Código de {c['nombre']} {c['apellido']}: {c['codigo_acceso']}", show_alert=True)
        return ConversationHandler.END

    elif accion.startswith("op_eliminar_"):
        op_id  = int(accion.replace("op_eliminar_", ""))
        r      = supabase.table("usuarios").select("nombre").eq("id", op_id).execute()
        nombre = r.data[0]["nombre"] if r.data else "este operario"
        if tiene_descargas_operario(op_id):
            teclado = InlineKeyboardMarkup([
                [InlineKeyboardButton("🚫 Quitar acceso al bot", callback_data=f"op_quitar_acceso_{op_id}")],
                [InlineKeyboardButton("❌ Cancelar",              callback_data=f"op_detalle_{op_id}")],
            ])
            await query.edit_message_text(
                f"⚠️ *{nombre}* tiene descargas registradas y no puede eliminarse.\n\n"
                "Podés quitarle el acceso al bot. Si en el futuro querés restaurar su acceso, "
                "se generará un nuevo código para compartirle.",
                parse_mode="Markdown", reply_markup=teclado
            )
        else:
            teclado = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Sí, eliminar", callback_data=f"op_confirmar_eliminar_{op_id}")],
                [InlineKeyboardButton("❌ Cancelar",      callback_data=f"op_detalle_{op_id}")],
            ])
            await query.edit_message_text(
                f"¿Confirmar eliminación de *{nombre}*?",
                parse_mode="Markdown", reply_markup=teclado
            )
        return ConversationHandler.END

    elif accion.startswith("op_confirmar_eliminar_"):
        op_id  = int(accion.replace("op_confirmar_eliminar_", ""))
        r      = supabase.table("usuarios").select("nombre").eq("id", op_id).execute()
        nombre = r.data[0]["nombre"] if r.data else ""
        supabase.table("usuarios").delete().eq("id", op_id).execute()
        await query.edit_message_text(
            f"✅ Operario *{nombre}* eliminado.",
            parse_mode="Markdown", reply_markup=teclado_menu_contratista(cont["id"], uid)
        )
        return ConversationHandler.END

    elif accion.startswith("op_quitar_acceso_"):
        op_id  = int(accion.replace("op_quitar_acceso_", ""))
        r      = supabase.table("usuarios").select("nombre").eq("id", op_id).execute()
        nombre = r.data[0]["nombre"] if r.data else ""
        nuevo_codigo = generar_codigo()
        supabase.table("usuarios").update({"telegram_id": None, "codigo_acceso": nuevo_codigo}).eq("id", op_id).execute()
        await query.edit_message_text(
            f"🚫 Acceso de *{nombre}* revocado.\n\n"
            f"Si querés restaurar su acceso, compartile este código: *{nuevo_codigo}*",
            parse_mode="Markdown", reply_markup=teclado_menu_contratista(cont["id"], uid)
        )
        return ConversationHandler.END

    elif accion.startswith("cli_eliminar_"):
        cli_id = int(accion.replace("cli_eliminar_", ""))
        r      = supabase.table("clientes").select("nombre,apellido").eq("id", cli_id).execute()
        nombre = f"{r.data[0]['nombre']} {r.data[0]['apellido']}" if r.data else "este cliente"
        if tiene_descargas_cliente(cli_id):
            teclado = InlineKeyboardMarkup([
                [InlineKeyboardButton("🚫 Quitar acceso al bot", callback_data=f"cli_quitar_acceso_{cli_id}")],
                [InlineKeyboardButton("❌ Cancelar",              callback_data=f"cli_detalle_{cli_id}")],
            ])
            await query.edit_message_text(
                f"⚠️ *{nombre}* tiene descargas registradas y no puede eliminarse.\n\n"
                "Podés quitarle el acceso al bot. Si en el futuro querés restaurar su acceso, "
                "se generará un nuevo código para compartirle.",
                parse_mode="Markdown", reply_markup=teclado
            )
        else:
            teclado = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Sí, eliminar", callback_data=f"cli_confirmar_eliminar_{cli_id}")],
                [InlineKeyboardButton("❌ Cancelar",      callback_data=f"cli_detalle_{cli_id}")],
            ])
            await query.edit_message_text(
                f"¿Confirmar eliminación de *{nombre}*?",
                parse_mode="Markdown", reply_markup=teclado
            )
        return ConversationHandler.END

    elif accion.startswith("cli_confirmar_eliminar_"):
        cli_id = int(accion.replace("cli_confirmar_eliminar_", ""))
        r      = supabase.table("clientes").select("nombre,apellido").eq("id", cli_id).execute()
        nombre = f"{r.data[0]['nombre']} {r.data[0]['apellido']}" if r.data else ""
        supabase.table("clientes").delete().eq("id", cli_id).execute()
        await query.edit_message_text(
            f"✅ Cliente *{nombre}* eliminado.",
            parse_mode="Markdown", reply_markup=teclado_menu_contratista(cont["id"], uid)
        )
        return ConversationHandler.END

    elif accion.startswith("cli_quitar_acceso_"):
        cli_id = int(accion.replace("cli_quitar_acceso_", ""))
        r      = supabase.table("clientes").select("nombre,apellido").eq("id", cli_id).execute()
        nombre = f"{r.data[0]['nombre']} {r.data[0]['apellido']}" if r.data else ""
        nuevo_codigo = generar_codigo()
        supabase.table("clientes").update({"telegram_id": None, "codigo_acceso": nuevo_codigo}).eq("id", cli_id).execute()
        await query.edit_message_text(
            f"🚫 Acceso de *{nombre}* revocado.\n\n"
            f"Si querés restaurar su acceso, compartile este código: *{nuevo_codigo}*",
            parse_mode="Markdown", reply_markup=teclado_menu_contratista(cont["id"], uid)
        )
        return ConversationHandler.END

    elif accion.startswith("op_editar_"):
        op_id = int(accion.replace("op_editar_", ""))
        context.user_data["editando_op_id"]  = op_id
        context.user_data["contratista_id"]  = cont["id"]
        await query.edit_message_text("¿Cuál es el nuevo nombre del operario?")
        return ADD_OP_NOMBRE

    elif accion.startswith("cli_editar_"):
        cli_id = int(accion.replace("cli_editar_", ""))
        context.user_data["editando_cli_id"] = cli_id
        context.user_data["contratista_id"]  = cont["id"]
        await query.edit_message_text("¿Cuál es el nuevo nombre del cliente?")
        return ADD_CLI_NOMBRE

    elif accion == "cont_add_op":
        context.user_data["contratista_id"] = cont["id"]
        await query.edit_message_text("¿Nombre y apellido del operario?")
        return ADD_OP_NOMBRE

    elif accion == "cont_add_cli":
        context.user_data["contratista_id"] = cont["id"]
        await query.edit_message_text("¿Nombre y apellido del cliente?")
        return ADD_CLI_NOMBRE

    elif accion == "cont_volver":
        await query.edit_message_text("¿Qué querés hacer?", reply_markup=teclado_menu_contratista(cont["id"], uid))
        return ConversationHandler.END

    # ── Mis descargas ─────────────────────────────────────────
    elif accion == "cont_ver_desc":
        clientes   = get_clientes(contratista_id)
        volver_cb  = "cont_volver" if cont else "op_menu"
        botones    = []
        for c in clientes:
            botones.append([InlineKeyboardButton(
                f"👤 {c['nombre']} {c['apellido']}", callback_data=f"cont_desc_cli_{c['id']}"
            )])
        if not clientes:
            botones.append([InlineKeyboardButton("(sin clientes)", callback_data=volver_cb)])
        botones.append([InlineKeyboardButton("📥 Exportar todo a Excel", callback_data="cont_desc_excel_todo")])
        botones.append([InlineKeyboardButton("⬅️ Volver", callback_data=volver_cb)])
        await query.edit_message_text(
            "📊 *Mis descargas*\n¿Para qué cliente?",
            parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones)
        )
        return ConversationHandler.END

    elif accion.startswith("cont_desc_cli_"):
        cli_id     = int(accion.replace("cont_desc_cli_", ""))
        r          = supabase.table("clientes").select("nombre,apellido").eq("id", cli_id).execute()
        cli_nombre = f"{r.data[0]['nombre']} {r.data[0]['apellido']}" if r.data else "Cliente"
        # Si el usuario logueado es el mismo cliente, volver a su menú propio
        cli_propio = get_cliente_by_telegram(uid)
        volver_cb  = "cli_ver_desc" if (cli_propio and cli_propio["id"] == cli_id) else "cont_ver_desc"
        campos  = get_campos(cli_id)
        botones = []
        for c in campos:
            botones.append([InlineKeyboardButton(f"🌿 {c['nombre']}", callback_data=f"cont_desc_campo_{c['id']}")])
        if not campos:
            botones.append([InlineKeyboardButton("(sin campos)", callback_data=volver_cb)])
        botones.append([InlineKeyboardButton("⬅️ Volver", callback_data=volver_cb)])
        await query.edit_message_text(
            f"📊 *{cli_nombre}*\n¿En qué campo?",
            parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones)
        )
        return ConversationHandler.END

    elif accion.startswith("cont_desc_campo_"):
        campo_id     = int(accion.replace("cont_desc_campo_", ""))
        r            = supabase.table("campos").select("nombre,cliente_id").eq("id", campo_id).execute()
        campo_nombre = r.data[0]["nombre"]    if r.data else "Campo"
        cli_id_back  = r.data[0]["cliente_id"] if r.data else 0
        lotes   = get_lotes(campo_id)
        botones = []
        for l in lotes:
            grano = f" ({l['grano']})" if l.get("grano") else ""
            botones.append([InlineKeyboardButton(f"📦 {l['nombre']}{grano}", callback_data=f"cont_desc_lote_{l['id']}")])
        if not lotes:
            botones.append([InlineKeyboardButton("(sin lotes)", callback_data=f"cont_desc_cli_{cli_id_back}")])
        botones.append([InlineKeyboardButton("⬅️ Volver", callback_data=f"cont_desc_cli_{cli_id_back}")])
        await query.edit_message_text(
            f"📊 *{campo_nombre}*\n¿En qué lote?",
            parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones)
        )
        return ConversationHandler.END

    elif accion.startswith("cont_desc_lote_") and not accion.startswith("cont_desc_lote_excel_"):
        lote_id  = int(accion.replace("cont_desc_lote_", ""))
        r_lote   = supabase.table("lotes").select("nombre,grano,campo_id").eq("id", lote_id).execute()
        if not r_lote.data:
            await query.edit_message_text("No se encontró el lote.")
            return ConversationHandler.END
        lote          = r_lote.data[0]
        campo_id_back = lote["campo_id"]
        descargas     = get_descargas_lote(lote_id)

        camiones_res = {}
        silos_res    = {}
        total_kg     = 0.0
        for d in descargas:
            kg = float(d.get("kg", 0))
            total_kg += kg
            if d.get("destino") == "camion" and d.get("camion_id"):
                cid = d["camion_id"]
                if cid not in camiones_res:
                    c = d.get("camiones") or {}
                    camiones_res[cid] = {
                        "patente": f"{c.get('patente_chasis','')} / {c.get('patente_acoplado','')}",
                        "kg": 0.0, "cerrado": c.get("cerrado", False)
                    }
                camiones_res[cid]["kg"] += kg
            elif d.get("destino") == "silo" and d.get("silobolsa_id"):
                sid = d["silobolsa_id"]
                if sid not in silos_res:
                    s = d.get("silobolsas") or {}
                    silos_res[sid] = {"numero": s.get("numero", "?"), "kg": 0.0, "cerrado": s.get("cerrado", False)}
                silos_res[sid]["kg"] += kg

        grano_str = f" ({lote['grano']})" if lote.get("grano") else ""
        lineas = [f"📊 *Lote: {lote['nombre']}{grano_str}*", f"Total: *{total_kg:,.0f} kg*"]
        if camiones_res:
            lineas.append("\n🚛 *Camiones:*")
            for c in camiones_res.values():
                estado = "🔒" if c["cerrado"] else "🟢"
                lineas.append(f"  {estado} {c['patente']} — {c['kg']:,.0f} kg")
        if silos_res:
            lineas.append("\n🌾 *Silobolsas:*")
            for s in silos_res.values():
                estado = "🔒" if s["cerrado"] else "🟢"
                lineas.append(f"  {estado} Silo #{s['numero']} — {s['kg']:,.0f} kg")
        if not camiones_res and not silos_res:
            lineas.append("\nSin descargas registradas.")

        botones = [
            [InlineKeyboardButton("📥 Exportar a Excel", callback_data=f"cont_desc_excel_{lote_id}")],
            [InlineKeyboardButton("⬅️ Volver",           callback_data=f"cont_desc_campo_{campo_id_back}")],
        ]
        await query.edit_message_text(
            "\n".join(lineas), parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones)
        )
        return ConversationHandler.END

    elif accion.startswith("cont_desc_excel_"):
        parte = accion.replace("cont_desc_excel_", "")
        await query.edit_message_text("⏳ Generando Excel, aguardá un momento...")
        if parte == "todo":
            descargas = get_descargas_contratista(contratista_id)
            filename  = "descargas_completo.xlsx"
            caption   = "📊 Reporte completo de descargas."
        else:
            lote_id   = int(parte)
            descargas = get_descargas_lote(lote_id)
            r_ln      = supabase.table("lotes").select("nombre").eq("id", lote_id).execute()
            lote_nombre = r_ln.data[0]["nombre"] if r_ln.data else "lote"
            filename  = f"descargas_{lote_nombre}.xlsx"
            caption   = f"📊 Reporte de descargas — Lote {lote_nombre}."
        buf = generar_excel(descargas)
        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=buf,
            filename=filename,
            caption=caption
        )
        menu_markup = (teclado_menu_contratista(contratista_id, uid) if cont
                       else teclado_menu_operario())
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="¿Qué querés hacer?",
            reply_markup=menu_markup
        )
        return ConversationHandler.END

    return ConversationHandler.END

# ── Agregar operario ─────────────────────────────────────────
async def add_op_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nombre         = update.message.text.strip()
    contratista_id = context.user_data["contratista_id"]
    uid            = str(update.effective_user.id)

    if "editando_op_id" in context.user_data:
        op_id = context.user_data.pop("editando_op_id")
        supabase.table("usuarios").update({"nombre": nombre.split()[0]}).eq("id", op_id).execute()
        await update.message.reply_text(f"✅ Nombre actualizado a *{nombre}*.", parse_mode="Markdown", reply_markup=teclado_menu_contratista(contratista_id, uid))
        return ConversationHandler.END

    context.user_data["nuevo_nombre"] = nombre
    teclado = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Sí, soy yo",         callback_data="op_soy_yo")],
        [InlineKeyboardButton("❌ No, es otra persona", callback_data="op_otro")],
    ])
    await update.message.reply_text(f"¿El operario *{nombre}* sos vos mismo?", parse_mode="Markdown", reply_markup=teclado)
    return ADD_OP_SOY_YO

async def add_op_soy_yo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query          = update.callback_query
    await query.answer()
    uid            = str(update.effective_user.id)
    nombre         = context.user_data["nuevo_nombre"]
    contratista_id = context.user_data["contratista_id"]
    partes         = nombre.split()

    if query.data == "op_soy_yo":
        r = supabase.table("usuarios").select("*").eq("telegram_id", uid).eq("contratista_id", contratista_id).execute()
        if r.data:
            await query.edit_message_text("Ya estás registrado como operario.", reply_markup=teclado_menu_contratista(contratista_id, uid))
            return ConversationHandler.END
        supabase.table("usuarios").insert({"nombre": partes[0], "rol": "operario", "telegram_id": uid, "contratista_id": contratista_id, "activo": True}).execute()
        await query.edit_message_text("✅ Quedaste registrado como operario.", reply_markup=teclado_menu_contratista(contratista_id, uid))
    else:
        codigo = generar_codigo()
        supabase.table("usuarios").insert({"nombre": partes[0], "rol": "operario", "codigo_acceso": codigo, "contratista_id": contratista_id, "activo": True}).execute()
        await query.edit_message_text(f"✅ Operario *{nombre}* creado.\n\nCódigo de acceso: *{codigo}*\nCompartíselo para que ingrese al bot.", parse_mode="Markdown", reply_markup=teclado_menu_contratista(contratista_id, uid))
    return ConversationHandler.END

# ── Agregar cliente ──────────────────────────────────────────
async def add_cli_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nombre         = update.message.text.strip()
    contratista_id = context.user_data["contratista_id"]
    uid            = str(update.effective_user.id)

    if "editando_cli_id" in context.user_data:
        cli_id = context.user_data.pop("editando_cli_id")
        partes  = nombre.split()
        supabase.table("clientes").update({"nombre": partes[0], "apellido": " ".join(partes[1:]) if len(partes) > 1 else ""}).eq("id", cli_id).execute()
        await update.message.reply_text(f"✅ Nombre actualizado a *{nombre}*.", parse_mode="Markdown", reply_markup=teclado_menu_contratista(contratista_id, uid))
        return ConversationHandler.END

    context.user_data["nuevo_nombre"] = nombre
    teclado = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Sí, soy yo",         callback_data="cli_soy_yo")],
        [InlineKeyboardButton("❌ No, es otra persona", callback_data="cli_otro")],
    ])
    await update.message.reply_text(f"¿El cliente *{nombre}* sos vos mismo?", parse_mode="Markdown", reply_markup=teclado)
    return ADD_CLI_SOY_YO

async def add_cli_soy_yo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query          = update.callback_query
    await query.answer()
    uid            = str(update.effective_user.id)
    nombre         = context.user_data["nuevo_nombre"]
    contratista_id = context.user_data["contratista_id"]
    partes         = nombre.split()
    nombre_p       = partes[0]
    apellido_p     = " ".join(partes[1:]) if len(partes) > 1 else ""

    if query.data == "cli_soy_yo":
        r = supabase.table("clientes").select("*").eq("telegram_id", uid).eq("contratista_id", contratista_id).execute()
        if r.data:
            await query.edit_message_text("Ya estás registrado como cliente.", reply_markup=teclado_menu_contratista(contratista_id, uid))
            return ConversationHandler.END
        supabase.table("clientes").insert({"nombre": nombre_p, "apellido": apellido_p, "telegram_id": uid, "contratista_id": contratista_id}).execute()
        await query.edit_message_text("✅ Quedaste registrado como cliente.", reply_markup=teclado_menu_contratista(contratista_id, uid))
    else:
        codigo = generar_codigo()
        supabase.table("clientes").insert({"nombre": nombre_p, "apellido": apellido_p, "codigo_acceso": codigo, "contratista_id": contratista_id}).execute()
        await query.edit_message_text(f"✅ Cliente *{nombre}* creado.\n\nCódigo de acceso: *{codigo}*\nCompartíselo para que ingrese al bot.", parse_mode="Markdown", reply_markup=teclado_menu_contratista(contratista_id, uid))
    return ConversationHandler.END

# ── Flujo de descarga ────────────────────────────────────────
async def iniciar_descarga(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    uid     = str(update.effective_user.id)
    chat_id = str(update.effective_chat.id)

    contratista_id = get_contratista_id_de_usuario(uid)
    if not contratista_id:
        await query.edit_message_text("No se encontró tu contratista. Escribí /start.")
        return ConversationHandler.END

    context.user_data["contratista_id"] = contratista_id
    sesion = get_sesion(chat_id)

    # Si el operario no tiene sesión propia, buscar la del contratista
    if not sesion:
        sesion = get_sesion_por_contratista(contratista_id)

    if sesion and sesion.get("lote_id"):
        cliente_obj = sesion.get("clientes") or {}
        campo_obj   = sesion.get("campos")   or {}
        lote_obj    = sesion.get("lotes")    or {}
        grano       = lote_obj.get("grano", "Sin definir")
        texto = (
            f"📋 *Sesión activa:*\n"
            f"Cliente: *{cliente_obj.get('nombre','')} {cliente_obj.get('apellido','')}*\n"
            f"Campo: *{campo_obj.get('nombre','')}*\n"
            f"Lote: *{lote_obj.get('nombre','')}* ({grano})"
        )
        teclado = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Continuar",      callback_data="desc_continuar")],
            [InlineKeyboardButton("🔄 Cambiar datos",  callback_data="desc_cambiar")],
            [InlineKeyboardButton("❌ Cancelar",        callback_data="op_cancelar")],
        ])
        await query.edit_message_text(texto, parse_mode="Markdown", reply_markup=teclado)
        return ConversationHandler.END
    else:
        return await mostrar_clientes(query, context, contratista_id, sesion)

async def mostrar_clientes(query, context, contratista_id, sesion=None):
    clientes = get_clientes(contratista_id)
    if not clientes:
        await query.edit_message_text(
            "No tenés clientes registrados. El contratista debe agregar clientes primero.",
            reply_markup=InlineKeyboardMarkup([[btn_cancelar()]]))
        return ConversationHandler.END

    sesion_cli_id = (sesion or {}).get("cliente_id")
    botones = []
    for c in clientes:
        marca = "✅ " if c["id"] == sesion_cli_id else ""
        botones.append([InlineKeyboardButton(f"{marca}{c['nombre']} {c['apellido']}", callback_data=f"desc_cli_{c['id']}")])
    botones.append([InlineKeyboardButton("➕ Nuevo cliente", callback_data="desc_nuevo_cliente")])
    botones.append([btn_cancelar()])
    await query.edit_message_text("¿Para qué cliente es esta cosecha?", reply_markup=InlineKeyboardMarkup(botones))
    return DESC_CLIENTE

async def desc_elegir_cliente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data  = query.data

    if data == "desc_nuevo_cliente":
        await query.edit_message_text("¿Nombre y apellido del nuevo cliente?")
        return NUEVO_CLIENTE_NOMBRE

    cli_id = int(data.replace("desc_cli_", ""))
    r      = supabase.table("clientes").select("*").eq("id", cli_id).execute()
    if not r.data:
        await query.edit_message_text("No se encontró el cliente.")
        return ConversationHandler.END
    c = r.data[0]
    context.user_data["desc_cliente_id"]  = cli_id
    context.user_data["desc_cliente_str"] = f"{c['nombre']} {c['apellido']}"
    return await mostrar_campos(query, context, cli_id)

async def desc_nuevo_cliente_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto          = update.message.text.strip()
    partes         = texto.split()
    contratista_id = context.user_data["contratista_id"]
    nuevo = supabase.table("clientes").insert({
        "nombre": partes[0], "apellido": " ".join(partes[1:]) if len(partes) > 1 else "",
        "contratista_id": contratista_id
    }).execute()
    cli_id = nuevo.data[0]["id"]
    context.user_data["desc_cliente_id"]  = cli_id
    context.user_data["desc_cliente_str"] = texto
    teclado = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Nuevo campo", callback_data="desc_nuevo_campo")],
        [btn_cancelar()]
    ])
    await update.message.reply_text(f"Cliente *{texto}* creado.\n\n¿En qué campo?", parse_mode="Markdown", reply_markup=teclado)
    return DESC_CAMPO

async def mostrar_campos(query, context, cliente_id):
    campos = get_campos(cliente_id)
    sesion_campo_id = context.user_data.get("desc_sesion_campo_id")
    botones = []
    for c in campos:
        marca = "✅ " if c["id"] == sesion_campo_id else ""
        botones.append([InlineKeyboardButton(f"{marca}{c['nombre']}", callback_data=f"desc_campo_{c['id']}")])
    botones.append([InlineKeyboardButton("➕ Nuevo campo", callback_data="desc_nuevo_campo")])
    botones.append([btn_cancelar()])
    await query.edit_message_text("¿En qué campo?", reply_markup=InlineKeyboardMarkup(botones))
    return DESC_CAMPO

async def desc_elegir_campo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data  = query.data

    if data == "desc_nuevo_campo":
        await query.edit_message_text("¿Nombre del nuevo campo?")
        return NUEVO_CAMPO_NOMBRE

    campo_id = int(data.replace("desc_campo_", ""))
    r        = supabase.table("campos").select("*").eq("id", campo_id).execute()
    if not r.data:
        await query.edit_message_text("No se encontró el campo.")
        return ConversationHandler.END
    context.user_data["desc_campo_id"]  = campo_id
    context.user_data["desc_campo_str"] = r.data[0]["nombre"]
    return await mostrar_lotes(query, context, campo_id)

async def desc_nuevo_campo_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto      = update.message.text.strip()
    cliente_id = context.user_data["desc_cliente_id"]
    nuevo      = supabase.table("campos").insert({"nombre": texto, "cliente_id": cliente_id}).execute()
    campo_id   = nuevo.data[0]["id"]
    context.user_data["desc_campo_id"]  = campo_id
    context.user_data["desc_campo_str"] = texto
    teclado = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Nuevo lote", callback_data="desc_nuevo_lote")],
        [btn_cancelar()]
    ])
    await update.message.reply_text(f"Campo *{texto}* creado.\n\n¿En qué lote?", parse_mode="Markdown", reply_markup=teclado)
    return DESC_LOTE

async def mostrar_lotes(query, context, campo_id):
    lotes           = get_lotes(campo_id)
    sesion_lote_id  = context.user_data.get("desc_sesion_lote_id")
    botones = []
    for l in lotes:
        marca = "✅ " if l["id"] == sesion_lote_id else ""
        grano = f" ({l['grano']})" if l.get("grano") else ""
        botones.append([InlineKeyboardButton(f"{marca}{l['nombre']}{grano}", callback_data=f"desc_lote_{l['id']}")])
    botones.append([InlineKeyboardButton("➕ Nuevo lote", callback_data="desc_nuevo_lote")])
    botones.append([btn_cancelar()])
    await query.edit_message_text("¿En qué lote?", reply_markup=InlineKeyboardMarkup(botones))
    return DESC_LOTE

async def desc_elegir_lote(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data  = query.data

    if data == "desc_nuevo_lote":
        await query.edit_message_text("¿Nombre del nuevo lote?")
        return NUEVO_LOTE_NOMBRE

    lote_id = int(data.replace("desc_lote_", ""))
    r       = supabase.table("lotes").select("*").eq("id", lote_id).execute()
    if not r.data:
        await query.edit_message_text("No se encontró el lote.")
        return ConversationHandler.END
    lote = r.data[0]
    context.user_data["desc_lote_id"]  = lote_id
    context.user_data["desc_lote_str"] = lote["nombre"]

    if not lote.get("grano"):
        return await mostrar_granos(query)

    context.user_data["desc_grano"] = lote["grano"]
    await _guardar_sesion(context, str(query.message.chat_id))
    return await mostrar_tipo_destino(query, context)

async def desc_nuevo_lote_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto    = update.message.text.strip()
    campo_id = context.user_data["desc_campo_id"]
    nuevo    = supabase.table("lotes").insert({"nombre": texto, "campo_id": campo_id}).execute()
    lote_id  = nuevo.data[0]["id"]
    context.user_data["desc_lote_id"]  = lote_id
    context.user_data["desc_lote_str"] = texto
    iconos  = {"Trigo": "🌾", "Soja": "🌱", "Maíz": "🌽", "Girasol": "🌻", "Sorgo": "🧅"}
    botones = []
    for g in GRANOS:
        botones.append([InlineKeyboardButton(f"{iconos.get(g,'')} {g}", callback_data=f"desc_grano_{g}")])
    botones.append([InlineKeyboardButton("✏️ Otro", callback_data="desc_grano_otro")])
    botones.append([btn_cancelar()])
    await update.message.reply_text(
        f"Lote *{texto}* creado.\n\n¿Qué grano se cosecha en este lote?",
        parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones)
    )
    return DESC_GRANO_OTRO

async def mostrar_granos(query):
    botones = []
    iconos  = {"Trigo": "🌾", "Soja": "🌱", "Maíz": "🌽", "Girasol": "🌻", "Sorgo": "🧅"}
    for g in GRANOS:
        botones.append([InlineKeyboardButton(f"{iconos.get(g,'')} {g}", callback_data=f"desc_grano_{g}")])
    botones.append([InlineKeyboardButton("✏️ Otro", callback_data="desc_grano_otro")])
    botones.append([btn_cancelar()])
    await query.edit_message_text("¿Qué grano se cosecha en este lote?", reply_markup=InlineKeyboardMarkup(botones))
    return DESC_LOTE

async def desc_elegir_grano(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data  = query.data

    if data == "desc_grano_otro":
        await query.edit_message_text("¿Cuál es el grano?")
        return DESC_GRANO_OTRO

    grano = data.replace("desc_grano_", "")
    context.user_data["desc_grano"] = grano
    supabase.table("lotes").update({"grano": grano}).eq("id", context.user_data["desc_lote_id"]).execute()
    await _guardar_sesion(context, str(query.message.chat_id))
    return await mostrar_tipo_destino(query, context)

async def desc_grano_otro(update: Update, context: ContextTypes.DEFAULT_TYPE):
    grano = update.message.text.strip()
    context.user_data["desc_grano"] = grano
    supabase.table("lotes").update({"grano": grano}).eq("id", context.user_data["desc_lote_id"]).execute()
    await _guardar_sesion(context, str(update.effective_chat.id))
    botones = [
        [InlineKeyboardButton("🚛 Camión",    callback_data="desc_tipo_camion")],
        [InlineKeyboardButton("🌾 Silobolsa", callback_data="desc_tipo_silo")],
        [btn_cancelar()]
    ]
    await update.message.reply_text("¿A dónde va la descarga?", reply_markup=InlineKeyboardMarkup(botones))
    return DESC_TIPO_DESTINO

async def _guardar_sesion(context, chat_id):
    supabase.table("sesion_activa").upsert({
        "chat_id":        chat_id,
        "contratista_id": context.user_data["contratista_id"],
        "cliente_id":     context.user_data["desc_cliente_id"],
        "campo_id":       context.user_data["desc_campo_id"],
        "lote_id":        context.user_data["desc_lote_id"],
        "iniciada_at":    ahora().isoformat()
    }).execute()

async def mostrar_tipo_destino(query, context):
    botones = [
        [InlineKeyboardButton("🚛 Camión",    callback_data="desc_tipo_camion")],
        [InlineKeyboardButton("🌾 Silobolsa", callback_data="desc_tipo_silo")],
        [btn_cancelar()]
    ]
    await query.edit_message_text("¿A dónde va la descarga?", reply_markup=InlineKeyboardMarkup(botones))
    return DESC_TIPO_DESTINO

async def desc_tipo_destino(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query          = update.callback_query
    await query.answer()
    contratista_id = context.user_data["contratista_id"]

    if query.data == "desc_tipo_camion":
        camiones = get_camiones_abiertos(contratista_id)
        botones  = []
        for c in camiones:
            cap_str = f" — faltan {c['capacidad_kg']-c['acumulado']:,.0f} kg" if c.get("capacidad_kg") else f" — {c['acumulado']:,.0f} kg acum."
            botones.append([InlineKeyboardButton(f"🚛 {c['patente_chasis']} / {c['patente_acoplado']}{cap_str}", callback_data=f"desc_cam_{c['id']}")])
        botones.append([InlineKeyboardButton("➕ Nuevo camión", callback_data="desc_nuevo_camion")])
        botones.append([btn_cancelar()])
        await query.edit_message_text("¿A qué camión?", reply_markup=InlineKeyboardMarkup(botones))
        context.user_data["desc_tipo"] = "camion"
        return DESC_TIPO_DESTINO

    else:
        silos   = get_silobolsas_abiertos(contratista_id)
        botones = []
        for s in silos:
            botones.append([InlineKeyboardButton(
                f"🌾 Silo #{s['numero']} — {s['lote_nombre']} ({s['grano']}) — {s['acumulado']:,.0f} kg",
                callback_data=f"desc_silo_{s['id']}"
            )])
        botones.append([InlineKeyboardButton("➕ Nuevo silobolsa", callback_data="desc_nuevo_silo")])
        botones.append([btn_cancelar()])
        await query.edit_message_text("¿A qué silobolsa?", reply_markup=InlineKeyboardMarkup(botones))
        context.user_data["desc_tipo"] = "silo"
        return DESC_TIPO_DESTINO

async def desc_elegir_destino(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data  = query.data

    if data == "desc_nuevo_camion":
        await query.edit_message_text("¿Patente del chasis? (ej: AB123CD)")
        return DESC_CAMION_CHASIS

    if data == "desc_nuevo_silo":
        sesion  = get_sesion(str(query.message.chat_id))
        lote_id = context.user_data.get("desc_lote_id") or (sesion["lote_id"] if sesion else None)
        if not lote_id:
            await query.edit_message_text("No se encontró el lote. Iniciá la sesión de nuevo.")
            return ConversationHandler.END
        r      = supabase.table("silobolsas").select("numero").eq("lote_id", lote_id).order("numero", desc=True).execute()
        numero = (r.data[0]["numero"] + 1) if r.data else 1
        nuevo  = supabase.table("silobolsas").insert({"numero": numero, "lote_id": lote_id}).execute()
        context.user_data["desc_destino_id"]  = nuevo.data[0]["id"]
        context.user_data["desc_destino_str"] = f"Silobolsa #{numero}"
        await query.edit_message_text(f"🌾 Silobolsa #{numero} creado.\n\n¿Cuántos kg?")
        return DESC_KG

    if data.startswith("desc_cam_"):
        camion_id = int(data.replace("desc_cam_", ""))
        r         = supabase.table("camiones").select("*").eq("id", camion_id).execute()
        camion    = r.data[0]
        context.user_data["desc_destino_id"]  = camion_id
        context.user_data["desc_destino_str"] = f"{camion['patente_chasis']} / {camion['patente_acoplado']}"
        context.user_data["desc_capacidad"]   = camion.get("capacidad_kg")
        await query.edit_message_text(f"🚛 {camion['patente_chasis']} / {camion['patente_acoplado']}\n\n¿Cuántos kg?")
        return DESC_KG

    if data.startswith("desc_silo_"):
        silo_id = int(data.replace("desc_silo_", ""))
        r       = supabase.table("silobolsas").select("numero").eq("id", silo_id).execute()
        numero  = r.data[0]["numero"] if r.data else "?"
        context.user_data["desc_destino_id"]  = silo_id
        context.user_data["desc_destino_str"] = f"Silobolsa #{numero}"
        await query.edit_message_text(f"🌾 Silobolsa #{numero}\n\n¿Cuántos kg?")
        return DESC_KG

    return ConversationHandler.END

async def desc_camion_chasis(update: Update, context: ContextTypes.DEFAULT_TYPE):
    patente = parsear_patente(update.message.text)
    r       = supabase.table("camiones").select("*").eq("patente_chasis", patente).execute()
    if r.data:
        camion = r.data[0]
        context.user_data["desc_chasis_tmp"] = patente
        context.user_data["desc_camion_tmp"] = camion
        teclado = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"✅ Sí, acoplado {camion['patente_acoplado']}", callback_data="desc_acoplado_ok")],
            [InlineKeyboardButton("❌ No, es otro acoplado",                        callback_data="desc_acoplado_otro")],
            [btn_cancelar()]
        ])
        cap_str = f"\nCapacidad: {camion['capacidad_kg']:,.0f} kg" if camion.get("capacidad_kg") else ""
        await update.message.reply_text(
            f"Encontré este camión:\nChasis: *{camion['patente_chasis']}*\nAcoplado: *{camion['patente_acoplado']}*{cap_str}\n\n¿Es este acoplado?",
            parse_mode="Markdown", reply_markup=teclado
        )
        return DESC_CAMION_ACOPLADO
    context.user_data["desc_chasis_tmp"] = patente
    await update.message.reply_text("¿Patente del acoplado?")
    return DESC_CAMION_ACOPLADO

async def desc_camion_acoplado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query if update.callback_query else None

    if query:
        await query.answer()
        if query.data == "desc_acoplado_ok":
            camion = context.user_data["desc_camion_tmp"]
            context.user_data["desc_destino_id"]  = camion["id"]
            context.user_data["desc_destino_str"] = f"{camion['patente_chasis']} / {camion['patente_acoplado']}"
            context.user_data["desc_capacidad"]   = camion.get("capacidad_kg")
            await query.edit_message_text(f"🚛 {camion['patente_chasis']} / {camion['patente_acoplado']}\n\n¿Cuántos kg?")
            return DESC_KG
        else:
            await query.edit_message_text("¿Patente del acoplado?")
            return DESC_CAMION_ACOPLADO
    else:
        patente = parsear_patente(update.message.text)
        context.user_data["desc_acoplado_tmp"] = patente
        await update.message.reply_text("¿Capacidad del camión en kg? (escribí 0 si no sabés)")
        return DESC_CAMION_CAPACIDAD

async def desc_camion_capacidad(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.strip().replace('.','').replace(',','.')
    try:
        cap = float(texto)
    except ValueError:
        cap = 0

    # Si es 0 o no ingresó nada → default 30000
    if cap == 0:
        cap = 30000
        context.user_data["desc_cap_tmp"] = cap
        await _guardar_camion_y_continuar(update, context, cap)
        await update.message.reply_text(
            f"No ingresaste capacidad, se asignaron *30.000 kg* por defecto.\n\n¿Cuántos kg descargaste?",
            parse_mode="Markdown"
        )
        return DESC_KG

    # Rango normal entre 25000 y 40000
    if 25000 <= cap <= 40000:
        await _guardar_camion_y_continuar(update, context, cap)
        return DESC_KG

    # Fuera de rango → preguntar si es correcto
    context.user_data["desc_cap_tmp"] = cap
    teclado = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Sí, es correcto", callback_data="cap_ok")],
        [InlineKeyboardButton("✏️ Corregir",        callback_data="cap_corregir")],
        [btn_cancelar()]
    ])
    await update.message.reply_text(
        f"Ingresaste *{cap:,.0f} kg* — eso parece inusual para un camión.\n¿Es correcto?",
        parse_mode="Markdown",
        reply_markup=teclado
    )
    return DESC_CAMION_CAPACIDAD

async def _guardar_camion_y_continuar(update, context, cap):
    chasis         = context.user_data["desc_chasis_tmp"]
    acoplado       = context.user_data["desc_acoplado_tmp"]
    contratista_id = context.user_data["contratista_id"]
    r = supabase.table("camiones").select("*").eq("patente_chasis", chasis).eq("patente_acoplado", acoplado).execute()
    if r.data:
        camion_id = r.data[0]["id"]
        supabase.table("camiones").update({"capacidad_kg": cap}).eq("id", camion_id).execute()
    else:
        nuevo     = supabase.table("camiones").insert({
            "patente_chasis": chasis, "patente_acoplado": acoplado,
            "capacidad_kg":   cap, "contratista_id": contratista_id
        }).execute()
        camion_id = nuevo.data[0]["id"]
    context.user_data["desc_destino_id"]  = camion_id
    context.user_data["desc_destino_str"] = f"{chasis} / {acoplado}"
    context.user_data["desc_capacidad"]   = cap
    await update.message.reply_text(
        f"🚛 Camión *{chasis} / {acoplado}* listo.\n\n¿Cuántos kg?",
        parse_mode="Markdown"
    )

async def desc_confirmar_capacidad(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "cap_corregir":
        await query.edit_message_text(
            "¿Cuál es la capacidad correcta en kg? (ej: 30000)"
        )
        return DESC_CAMION_CAPACIDAD

    # Confirma la capacidad inusual
    cap = context.user_data["desc_cap_tmp"]
    chasis         = context.user_data["desc_chasis_tmp"]
    acoplado       = context.user_data["desc_acoplado_tmp"]
    contratista_id = context.user_data["contratista_id"]
    r = supabase.table("camiones").select("*").eq("patente_chasis", chasis).eq("patente_acoplado", acoplado).execute()
    if r.data:
        camion_id = r.data[0]["id"]
        supabase.table("camiones").update({"capacidad_kg": cap}).eq("id", camion_id).execute()
    else:
        nuevo     = supabase.table("camiones").insert({
            "patente_chasis": chasis, "patente_acoplado": acoplado,
            "capacidad_kg":   cap, "contratista_id": contratista_id
        }).execute()
        camion_id = nuevo.data[0]["id"]
    context.user_data["desc_destino_id"]  = camion_id
    context.user_data["desc_destino_str"] = f"{chasis} / {acoplado}"
    context.user_data["desc_capacidad"]   = cap
    await query.edit_message_text(
        f"🚛 Camión *{chasis} / {acoplado}* listo.\n\n¿Cuántos kg?",
        parse_mode="Markdown"
    )
    return DESC_KG

async def desc_recibir_kg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kg = parsear_kg(update.message.text)
    if not kg:
        await update.message.reply_text("No entendí los kg. Escribí solo el número, ej: `13500`", parse_mode="Markdown")
        return DESC_KG

    context.user_data["desc_kg"] = kg
    chat_id = str(update.effective_chat.id)
    sesion  = get_sesion(chat_id)

    cliente_str = context.user_data.get("desc_cliente_str") or ""
    campo_str   = context.user_data.get("desc_campo_str")   or ""
    lote_str    = context.user_data.get("desc_lote_str")    or ""
    grano       = context.user_data.get("desc_grano")       or ""
    destino_str = context.user_data.get("desc_destino_str") or ""
    tipo        = context.user_data.get("desc_tipo", "camion")
    icono       = "🚛" if tipo == "camion" else "🌾"

    if not cliente_str and sesion:
        c = sesion.get("clientes") or {}
        cliente_str = f"{c.get('nombre','')} {c.get('apellido','')}".strip()
        campo_str   = (sesion.get("campos") or {}).get("nombre","")
        lote_str    = (sesion.get("lotes")  or {}).get("nombre","")
        grano       = (sesion.get("lotes")  or {}).get("grano","")

    texto = (
        f"📋 *Confirmar descarga*\n\n"
        f"Cliente: *{cliente_str}*\n"
        f"Campo: *{campo_str}*\n"
        f"Lote: *{lote_str}* ({grano})\n"
        f"Destino: {icono} *{destino_str}*\n"
        f"Kg: *{kg:,.0f}*"
    )
    teclado = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Confirmar",       callback_data="desc_confirmar")],
        [InlineKeyboardButton("👤 Cambiar cliente", callback_data="desc_cambiar_cliente"),
         InlineKeyboardButton("🌾 Cambiar campo",   callback_data="desc_cambiar_campo")],
        [InlineKeyboardButton("🌱 Cambiar lote",    callback_data="desc_cambiar_lote"),
         InlineKeyboardButton("🚛 Cambiar destino", callback_data="desc_cambiar_destino")],
        [InlineKeyboardButton("⚖️ Cambiar kg",      callback_data="desc_cambiar_kg")],
        [btn_cancelar()]
    ])
    await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=teclado)
    return DESC_KG

async def desc_confirmar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    uid     = str(update.effective_user.id)
    chat_id = str(update.effective_chat.id)
    data    = query.data

    if data == "desc_cambiar_kg":
        await query.edit_message_text("¿Cuántos kg?")
        return DESC_KG

    if data == "desc_cambiar_cliente":
        sesion = get_sesion(chat_id)
        context.user_data["desc_sesion_campo_id"] = sesion["campo_id"] if sesion else None
        context.user_data["desc_sesion_lote_id"]  = sesion["lote_id"]  if sesion else None
        return await mostrar_clientes(query, context, context.user_data["contratista_id"], sesion)

    if data == "desc_cambiar_campo":
        return await mostrar_campos(query, context, context.user_data["desc_cliente_id"])

    if data == "desc_cambiar_lote":
        return await mostrar_lotes(query, context, context.user_data["desc_campo_id"])

    if data == "desc_cambiar_destino":
        return await mostrar_tipo_destino(query, context)

    if data != "desc_confirmar":
        return DESC_KG

    # Guardar descarga
    kg             = context.user_data["desc_kg"]
    destino_id     = context.user_data["desc_destino_id"]
    tipo           = context.user_data.get("desc_tipo", "camion")
    sesion         = get_sesion(chat_id)
    operario       = get_usuario(uid)
    contratista_id = context.user_data["contratista_id"]

    lote_id    = context.user_data.get("desc_lote_id")   or (sesion["lote_id"]   if sesion else None)
    campo_id   = context.user_data.get("desc_campo_id")  or (sesion["campo_id"]  if sesion else None)
    cliente_id = context.user_data.get("desc_cliente_id") or (sesion["cliente_id"] if sesion else None)

    supabase.table("descargas").insert({
        "kg":              kg,
        "destino":         tipo,
        "camion_id":       destino_id if tipo == "camion" else None,
        "silobolsa_id":    destino_id if tipo == "silo"   else None,
        "lote_id":         lote_id,
        "campo_id":        campo_id,
        "cliente_id":      cliente_id,
        "contratista_id":  contratista_id,
        "operario_id":     operario["id"] if operario else None,
        "chat_id":         chat_id,
        "created_at":      ahora().isoformat(),
    }).execute()

    # Calcular acumulado y armar respuesta
    if tipo == "camion":
        acumulado = kg_acumulado_camion(destino_id)
        capacidad = context.user_data.get("desc_capacidad")
        destino_str = context.user_data.get("desc_destino_str","")
        lineas = [
            f"✅ Descarga registrada",
            f"🚛 *{destino_str}*",
            f"Esta carga:  *{kg:,.0f} kg*",
            f"Acumulado:   *{acumulado:,.0f} kg*" + (f" / {capacidad:,.0f} kg" if capacidad else ""),
        ]
        if capacidad:
            pct = acumulado / capacidad * 100
            lineas.append(barra(acumulado, capacidad))
            if pct >= 95:
                lineas.append(f"⚠️ {pct:.0f}% — ¿Se completó el camión?")
                teclado = InlineKeyboardMarkup([
                    [InlineKeyboardButton("✅ Sí, se fue lleno", callback_data=f"cam_cerrar_{destino_id}")],
                    [InlineKeyboardButton("❌ No, sigue cargando", callback_data="op_menu")],
                ])
                await query.edit_message_text("\n".join(lineas), parse_mode="Markdown", reply_markup=teclado)
                return ConversationHandler.END
    else:
        acumulado   = kg_acumulado_silo(destino_id)
        destino_str = context.user_data.get("desc_destino_str","")
        lineas = [
            f"✅ Descarga registrada",
            f"🌾 *{destino_str}*",
            f"Esta carga:  *{kg:,.0f} kg*",
            f"Acumulado:   *{acumulado:,.0f} kg*",
        ]

    uid_op = str(update.effective_user.id)
    cont   = get_contratista(uid_op)
    if cont:
        teclado = teclado_menu_contratista(cont["id"], uid_op)
    else:
        teclado = teclado_menu_operario()

    await query.edit_message_text("\n".join(lineas), parse_mode="Markdown", reply_markup=teclado)
    return ConversationHandler.END

# ── Menú operario callbacks ──────────────────────────────────
async def menu_operario_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    uid     = str(update.effective_user.id)
    accion  = query.data

    if accion == "op_menu":
        cont = get_contratista(uid)
        if cont:
            await query.edit_message_text("¿Qué querés hacer?", reply_markup=teclado_menu_contratista(cont["id"], uid))
        else:
            await query.edit_message_text("¿Qué querés hacer?", reply_markup=teclado_menu_operario())
        return ConversationHandler.END

    if accion == "op_cancelar":
        cont = get_contratista(uid)
        if cont:
            await query.edit_message_text("Operación cancelada.", reply_markup=teclado_menu_contratista(cont["id"], uid))
        else:
            await query.edit_message_text("Operación cancelada.", reply_markup=teclado_menu_operario())
        return ConversationHandler.END

    contratista_id = get_contratista_id_de_usuario(uid)

    if accion == "op_camiones":
        camiones = get_camiones_abiertos(contratista_id)
        r_cerrados = supabase.table("camiones").select("*").eq("contratista_id", contratista_id).eq("cerrado", True).execute()
        cerrados   = r_cerrados.data or []
        botones    = []
        lineas     = ["🚛 *Camiones*\n"]
        for c in camiones:
            cap_str = f" ({c['acumulado']:,.0f}/{c['capacidad_kg']:,.0f} kg)" if c.get("capacidad_kg") else f" ({c['acumulado']:,.0f} kg)"
            lineas.append(f"🟢 {c['patente_chasis']} / {c['patente_acoplado']}{cap_str}")
            botones.append([InlineKeyboardButton(f"🟢 {c['patente_chasis']} / {c['patente_acoplado']}", callback_data=f"cam_detalle_{c['id']}")])
        for c in cerrados:
            acum = kg_acumulado_camion(c["id"])
            lineas.append(f"🔒 {c['patente_chasis']} / {c['patente_acoplado']} ({acum:,.0f} kg)")
            botones.append([InlineKeyboardButton(f"🔒 {c['patente_chasis']} / {c['patente_acoplado']}", callback_data=f"cam_detalle_{c['id']}")])
        if not camiones and not cerrados:
            lineas = ["No hay camiones registrados todavía."]
        botones.append([btn_cancelar("op_menu")])
        await query.edit_message_text("\n".join(lineas), parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))
        return ConversationHandler.END

    if accion == "op_silos":
        silos_abiertos = get_silobolsas_abiertos(contratista_id)
        r_cerrados     = supabase.table("silobolsas").select("*, lotes(nombre,grano)").eq("cerrado", True).execute()
        cerrados       = r_cerrados.data or []
        botones        = []
        lineas         = ["🌾 *Silobolsas*\n"]
        for s in silos_abiertos:
            lineas.append(f"🟢 Silo #{s['numero']} — {s['lote_nombre']} ({s['grano']}) — {s['acumulado']:,.0f} kg")
            botones.append([InlineKeyboardButton(f"🟢 Silo #{s['numero']} — {s['lote_nombre']}", callback_data=f"silo_detalle_{s['id']}")])
        for s in cerrados:
            lote  = s.get("lotes") or {}
            acum  = kg_acumulado_silo(s["id"])
            lineas.append(f"🔒 Silo #{s['numero']} — {lote.get('nombre','')} — {acum:,.0f} kg")
            botones.append([InlineKeyboardButton(f"🔒 Silo #{s['numero']} — {lote.get('nombre','')}", callback_data=f"silo_detalle_{s['id']}")])
        if not silos_abiertos and not cerrados:
            lineas = ["No hay silobolsas registrados todavía."]
        botones.append([btn_cancelar("op_menu")])
        await query.edit_message_text("\n".join(lineas), parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))
        return ConversationHandler.END

    if accion.startswith("cam_detalle_"):
        camion_id = int(accion.replace("cam_detalle_", ""))
        r         = supabase.table("camiones").select("*").eq("id", camion_id).execute()
        if not r.data:
            await query.edit_message_text("No se encontró el camión.")
            return ConversationHandler.END
        camion    = r.data[0]
        acumulado = kg_acumulado_camion(camion_id)
        capacidad = camion.get("capacidad_kg")
        cerrado   = camion.get("cerrado", False)
        lineas    = [f"🚛 *{camion['patente_chasis']} / {camion['patente_acoplado']}*"]
        lineas.append(f"Acumulado: *{acumulado:,.0f} kg*" + (f" / {capacidad:,.0f} kg" if capacidad else ""))
        if capacidad: lineas.append(barra(acumulado, capacidad))
        if cerrado: lineas.append("Estado: 🔒 Cerrado")
        botones = []
        if cerrado:
            botones.append([InlineKeyboardButton("📦 Agregar descarga igual", callback_data=f"cam_forzar_desc_{camion_id}")])
            botones.append([InlineKeyboardButton("🔓 Reabrir",                callback_data=f"cam_reabrir_{camion_id}")])
        else:
            botones.append([InlineKeyboardButton("📦 Agregar descarga",       callback_data=f"cam_forzar_desc_{camion_id}")])
            botones.append([InlineKeyboardButton("🔒 Cerrar camión",          callback_data=f"cam_cerrar_{camion_id}")])
        botones.append([btn_cancelar("op_camiones")])
        await query.edit_message_text("\n".join(lineas), parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))
        return ConversationHandler.END

    if accion.startswith("silo_detalle_"):
        silo_id = int(accion.replace("silo_detalle_", ""))
        r       = supabase.table("silobolsas").select("*, lotes(nombre,grano)").eq("id", silo_id).execute()
        if not r.data:
            await query.edit_message_text("No se encontró el silobolsa.")
            return ConversationHandler.END
        silo    = r.data[0]
        lote    = silo.get("lotes") or {}
        acum    = kg_acumulado_silo(silo_id)
        cerrado = silo.get("cerrado", False)
        lineas  = [
            f"🌾 *Silobolsa #{silo['numero']}*",
            f"Lote: {lote.get('nombre','')} ({lote.get('grano','')})",
            f"Acumulado: *{acum:,.0f} kg*",
        ]
        if cerrado: lineas.append("Estado: 🔒 Cerrado")
        botones = []
        if cerrado:
            botones.append([InlineKeyboardButton("📦 Agregar descarga igual", callback_data=f"silo_forzar_desc_{silo_id}")])
            botones.append([InlineKeyboardButton("🔓 Reabrir",                callback_data=f"silo_reabrir_{silo_id}")])
        else:
            botones.append([InlineKeyboardButton("📦 Agregar descarga",       callback_data=f"silo_forzar_desc_{silo_id}")])
            botones.append([InlineKeyboardButton("🔒 Cerrar silobolsa",       callback_data=f"silo_cerrar_{silo_id}")])
        botones.append([btn_cancelar("op_silos")])
        await query.edit_message_text("\n".join(lineas), parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(botones))
        return ConversationHandler.END

    if accion.startswith("cam_cerrar_"):
        camion_id = int(accion.replace("cam_cerrar_", ""))
        supabase.table("camiones").update({"cerrado": True}).eq("id", camion_id).execute()
        await query.edit_message_text("🔒 Camión cerrado.", reply_markup=teclado_menu_operario())
        return ConversationHandler.END

    if accion.startswith("cam_reabrir_"):
        camion_id = int(accion.replace("cam_reabrir_", ""))
        supabase.table("camiones").update({"cerrado": False}).eq("id", camion_id).execute()
        await query.edit_message_text("🔓 Camión reabierto.", reply_markup=teclado_menu_operario())
        return ConversationHandler.END

    if accion.startswith("silo_cerrar_"):
        silo_id = int(accion.replace("silo_cerrar_", ""))
        supabase.table("silobolsas").update({"cerrado": True}).eq("id", silo_id).execute()
        await query.edit_message_text("🔒 Silobolsa cerrado.", reply_markup=teclado_menu_operario())
        return ConversationHandler.END

    if accion.startswith("silo_reabrir_"):
        silo_id = int(accion.replace("silo_reabrir_", ""))
        supabase.table("silobolsas").update({"cerrado": False}).eq("id", silo_id).execute()
        await query.edit_message_text("🔓 Silobolsa reabierto.", reply_markup=teclado_menu_operario())
        return ConversationHandler.END

    if accion.startswith("cam_forzar_desc_"):
        camion_id = int(accion.replace("cam_forzar_desc_", ""))
        r         = supabase.table("camiones").select("*").eq("id", camion_id).execute()
        camion    = r.data[0]
        context.user_data["desc_tipo"]         = "camion"
        context.user_data["desc_destino_id"]   = camion_id
        context.user_data["desc_destino_str"]  = f"{camion['patente_chasis']} / {camion['patente_acoplado']}"
        context.user_data["desc_capacidad"]    = camion.get("capacidad_kg")
        context.user_data["contratista_id"]    = get_contratista_id_de_usuario(uid)
        await query.edit_message_text(f"🚛 {camion['patente_chasis']} / {camion['patente_acoplado']}\n\n¿Cuántos kg?")
        return DESC_KG

    if accion.startswith("silo_forzar_desc_"):
        silo_id = int(accion.replace("silo_forzar_desc_", ""))
        r       = supabase.table("silobolsas").select("numero").eq("id", silo_id).execute()
        numero  = r.data[0]["numero"] if r.data else "?"
        context.user_data["desc_tipo"]        = "silo"
        context.user_data["desc_destino_id"]  = silo_id
        context.user_data["desc_destino_str"] = f"Silobolsa #{numero}"
        context.user_data["contratista_id"]   = get_contratista_id_de_usuario(uid)
        await query.edit_message_text(f"🌾 Silobolsa #{numero}\n\n¿Cuántos kg?")
        return DESC_KG

    if accion == "desc_continuar":
        sesion = get_sesion(str(query.message.chat_id))
        if sesion:
            context.user_data["contratista_id"]   = sesion["contratista_id"]
            context.user_data["desc_cliente_id"]  = sesion["cliente_id"]
            context.user_data["desc_campo_id"]    = sesion["campo_id"]
            context.user_data["desc_lote_id"]     = sesion["lote_id"]
            lote = sesion.get("lotes") or {}
            context.user_data["desc_grano"]       = lote.get("grano","")
            c = sesion.get("clientes") or {}
            context.user_data["desc_cliente_str"] = f"{c.get('nombre','')} {c.get('apellido','')}".strip()
            context.user_data["desc_campo_str"]   = (sesion.get("campos") or {}).get("nombre","")
            context.user_data["desc_lote_str"]    = lote.get("nombre","")
        return await mostrar_tipo_destino(query, context)

    if accion == "desc_cambiar":
        sesion         = get_sesion(str(query.message.chat_id))
        contratista_id = get_contratista_id_de_usuario(uid)
        context.user_data["contratista_id"]       = contratista_id
        context.user_data["desc_sesion_campo_id"] = sesion["campo_id"] if sesion else None
        context.user_data["desc_sesion_lote_id"]  = sesion["lote_id"]  if sesion else None
        return await mostrar_clientes(query, context, contratista_id, sesion)

    return ConversationHandler.END

# ── Handler mensajes libres ──────────────────────────────────
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = str(update.effective_user.id)
    cont = get_contratista(uid)
    if cont:
        await update.message.reply_text(f"Hola {cont['nombre']}! ¿Qué querés hacer?", reply_markup=teclado_menu_contratista(cont["id"], uid))
        return
    usr = get_usuario(uid)
    if usr:
        await update.message.reply_text(f"Hola {usr['nombre']}! ¿Qué querés hacer?", reply_markup=teclado_menu_operario())
        return
    cli = get_cliente_by_telegram(uid)
    if cli:
        cont_nombre = (cli.get("contratistas") or {}).get("nombre", "")
        await update.message.reply_text(
            f"Hola {cli['nombre']} {cli['apellido']}! Sos cliente de *{cont_nombre}*.\n¿Qué querés hacer?",
            parse_mode="Markdown", reply_markup=teclado_menu_cliente()
        )
        return
    await update.message.reply_text(
        "No tenés permisos para acceder.\n\nSi sos contratista escribí /start.\nSi sos operario o cliente, ingresá el código que te dio tu contratista:"
    )

# ── Main ─────────────────────────────────────────────────────
if __name__ == "__main__":
    app = ApplicationBuilder().token(TOKEN).build()

    registro_conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", cmd_start),
            CallbackQueryHandler(elegir_rol, pattern="^rol_"),
        ],
        states={
            REG_ROL:          [CallbackQueryHandler(elegir_rol, pattern="^rol_")],
            REG_NOMBRE:       [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_nombre_contratista)],
            INGRESAR_CODIGO:  [MessageHandler(filters.TEXT & ~filters.COMMAND, ingresar_codigo)],
            CONFIRMAR_NOMBRE: [CallbackQueryHandler(confirmar_nombre, pattern="^nombre_")],
            CORREGIR_NOMBRE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, corregir_nombre)],
        },
        fallbacks=[CommandHandler("start", cmd_start)],
        per_message=False
    )

    menu_cont_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(menu_contratista_callback, pattern="^(cont_|cli_ver_desc|cli_menu|op_detalle_|cli_detalle_|op_vercodigo_|cli_vercodigo_|op_eliminar_|cli_eliminar_|op_confirmar_eliminar_|cli_confirmar_eliminar_|op_quitar_acceso_|cli_quitar_acceso_|op_editar_|cli_editar_)")],
        states={
            ADD_OP_NOMBRE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, add_op_nombre)],
            ADD_OP_SOY_YO:  [CallbackQueryHandler(add_op_soy_yo,  pattern="^(op_soy_yo|op_otro)$")],
            ADD_CLI_NOMBRE: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_cli_nombre)],
            ADD_CLI_SOY_YO: [CallbackQueryHandler(add_cli_soy_yo, pattern="^(cli_soy_yo|cli_otro)$")],
        },
        fallbacks=[CommandHandler("start", cmd_start)],
        allow_reentry=True,
        per_message=False
    )

    descarga_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(iniciar_descarga, pattern="^op_descarga$"),
            CallbackQueryHandler(menu_operario_callback, pattern="^(desc_continuar$|desc_cambiar$)"),
        ],
        states={
            DESC_CLIENTE:       [CallbackQueryHandler(desc_elegir_cliente,  pattern="^desc_cli_|^desc_nuevo_cliente$")],
            NUEVO_CLIENTE_NOMBRE: [MessageHandler(filters.TEXT & ~filters.COMMAND, desc_nuevo_cliente_nombre)],
            DESC_CAMPO:         [CallbackQueryHandler(desc_elegir_campo,    pattern="^desc_campo_|^desc_nuevo_campo$")],
            NUEVO_CAMPO_NOMBRE: [MessageHandler(filters.TEXT & ~filters.COMMAND, desc_nuevo_campo_nombre)],
            DESC_LOTE:          [CallbackQueryHandler(desc_elegir_lote,     pattern="^desc_lote_|^desc_nuevo_lote$"),
                                 CallbackQueryHandler(desc_elegir_grano,    pattern="^desc_grano_")],
            NUEVO_LOTE_NOMBRE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, desc_nuevo_lote_nombre),
                                 CallbackQueryHandler(desc_elegir_grano, pattern="^desc_grano_")],
            DESC_GRANO_OTRO:    [MessageHandler(filters.TEXT & ~filters.COMMAND, desc_grano_otro),
                                 CallbackQueryHandler(desc_elegir_grano, pattern="^desc_grano_")],
            DESC_TIPO_DESTINO:  [CallbackQueryHandler(desc_tipo_destino,    pattern="^desc_tipo_"),
                                 CallbackQueryHandler(desc_elegir_destino,  pattern="^desc_cam_|^desc_silo_|^desc_nuevo_camion$|^desc_nuevo_silo$")],
            DESC_CAMION_CHASIS:   [MessageHandler(filters.TEXT & ~filters.COMMAND, desc_camion_chasis)],
            DESC_CAMION_ACOPLADO: [CallbackQueryHandler(desc_camion_acoplado, pattern="^desc_acoplado_"),
                                   MessageHandler(filters.TEXT & ~filters.COMMAND, desc_camion_acoplado)],
            DESC_CAMION_CAPACIDAD:[
                MessageHandler(filters.TEXT & ~filters.COMMAND, desc_camion_capacidad),
                CallbackQueryHandler(desc_confirmar_capacidad, pattern="^cap_"),
            ],
            DESC_KG:            [MessageHandler(filters.TEXT & ~filters.COMMAND, desc_recibir_kg),
                                 CallbackQueryHandler(desc_confirmar, pattern="^desc_confirmar$|^desc_cambiar_|^op_cancelar$")],
        },
        fallbacks=[
            CallbackQueryHandler(menu_operario_callback, pattern="^op_cancelar$"),
            CommandHandler("start", cmd_start)
        ],
        per_message=False
    )

    operario_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(menu_operario_callback, pattern="^(op_menu$|op_cancelar$|op_camiones$|op_silos$|cam_detalle_|silo_detalle_|cam_cerrar_|cam_reabrir_|silo_cerrar_|silo_reabrir_|cam_forzar_desc_|silo_forzar_desc_|desc_continuar$|desc_cambiar$)")],
        states={
            DESC_KG: [MessageHandler(filters.TEXT & ~filters.COMMAND, desc_recibir_kg),
                      CallbackQueryHandler(desc_confirmar, pattern="^desc_confirmar$|^desc_cambiar_|^op_cancelar$|^cam_cerrar_")],
        },
        fallbacks=[CommandHandler("start", cmd_start)],
        per_message=False
    )

    app.add_handler(registro_conv)
    app.add_handler(menu_cont_conv)
    app.add_handler(descarga_conv)
    app.add_handler(operario_conv)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    print("Bot corriendo...")
    app.run_polling()
